
# from flask import Flask, request, jsonify
# from openpyxl import Workbook, load_workbook
# from flask_cors import CORS
# import os

# app = Flask(__name__)
# CORS(app)  # Enable CORS for all routes

# # Define the file path
# file_path = 'rtsp_data.xlsx'

# def initialize_workbook():
#     # Create the file if it doesn't exist
#     if not os.path.exists(file_path):
#         try:
#             workbook = Workbook()
#             sheet = workbook.active
#             sheet.title = 'RTSP Data'
#             sheet.append(['RTSP URL'])
#             workbook.save(file_path)
#             print(f"Workbook created and saved as {file_path}")
#         except Exception as e:
#             print(f"Error initializing workbook: {e}")
#             raise

# @app.route('/rtsp', methods=['POST'])
# def save_rtsp_url():
#     data = request.json
#     rtsp_url = data.get('rtsp_url')

#     if rtsp_url:
#         try:
#             print(f"Received RTSP URL: {rtsp_url}")
#             initialize_workbook()
#             workbook = load_workbook(file_path)
#             sheet = workbook.active
#             sheet.append([rtsp_url])
#             workbook.save(file_path)
#             print(f"RTSP URL saved to {file_path}")
#             return jsonify({'status': 'success', 'data': {'rtsp URL': rtsp_url}})
#         except Exception as e:
#             print(f"Error saving RTSP URL: {e}")
#             return jsonify({'status': 'error', 'message': 'Failed to save RTSP URL'}), 500
#     return jsonify({'status': 'error', 'message': 'RTSP URL is required'}), 400

# if __name__ == '__main__':
#     app.run(debug=True, port=5000)











from flask import Flask, request, jsonify, redirect
from openpyxl import Workbook, load_workbook
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration for your AWS S3
AWS_BUCKET_NAME = 'rtspcamera'
AWS_REGION = 'ap-south-1'  # e.g., 'us-east-1'

# Define the file path
file_path = 'rtsp_data.xlsx'

def initialize_workbook():
    # Create the file if it doesn't exist
    if not os.path.exists(file_path):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'RTSP Data'
            sheet.append(['RTSP URL'])
            workbook.save(file_path)
            print(f"Workbook created and saved as {file_path}")
        except Exception as e:
            print(f"Error initializing workbook: {e}")
            raise

@app.route('/redirect-to-s3')
def redirect_to_s3():
    s3_url = f'https://s3.console.aws.amazon.com/s3/buckets/{AWS_BUCKET_NAME}?region={AWS_REGION}&tab=objects'
    return redirect(s3_url)

@app.route('/rtsp', methods=['POST'])
def save_rtsp_url():
    data = request.json
    rtsp_url = data.get('rtsp_url')

    if rtsp_url:
        try:
            print(f"Received RTSP URL: {rtsp_url}")
            initialize_workbook()
            workbook = load_workbook(file_path)
            sheet = workbook.active
            sheet.append([rtsp_url])
            workbook.save(file_path)
            print(f"RTSP URL saved to {file_path}")
            return jsonify({'status': 'success', 'data': {'rtsp URL': rtsp_url}})
        except Exception as e:
            print(f"Error saving RTSP URL: {e}")
            return jsonify({'status': 'error', 'message': 'Failed to save RTSP URL'}), 500
    return jsonify({'status': 'error', 'message': 'RTSP URL is required'}), 400

if __name__ == '__main__':
    app.run(debug=True, port=5000)
