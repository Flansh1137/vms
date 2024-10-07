from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Define the file path
file_path = 'cloud_data.xlsx'

def initialize_workbook():
    # Create the file if it doesn't exist
    if not os.path.exists(file_path):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Cloud Configurations'
            sheet.append(['Access Key', 'Secret Key', 'Bucket Name', 'Region'])
            workbook.save(file_path)
            print(f"Workbook created and saved as {file_path}")
        except Exception as e:
            print(f"Error initializing workbook: {e}")
            raise

@app.route('/submit-cloud', methods=['POST'])
def save_cloud_data():
    data = request.json
    access_key = data.get('access_key')
    secret_key = data.get('secret_key')
    bucket_name = data.get('bucket_name')
    region = data.get('region')

    if access_key and secret_key and bucket_name and region:
        try:
            print(f"Received cloud data: Access Key: {access_key}, Secret Key: {secret_key}, Bucket Name: {bucket_name}, Region: {region}")
            initialize_workbook()
            workbook = load_workbook(file_path)
            sheet = workbook.active
            sheet.append([access_key, secret_key, bucket_name, region])
            workbook.save(file_path)
            print(f"Cloud data saved to {file_path}")
            return jsonify({'status': 'success', 'data': {'access_key': access_key}})
        except Exception as e:
            print(f"Error saving cloud data: {e}")
            return jsonify({'status': 'error', 'message': 'Failed to save cloud data'}), 500
    return jsonify({'status': 'error', 'message': 'All fields are required'}), 400

if __name__ == '__main__':
    app.run(debug=True, port=5000)
